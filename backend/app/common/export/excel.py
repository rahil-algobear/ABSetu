"""
Generic XLSX writer for list/report exports.

`build_xlsx` takes plain headers + a matrix of already-resolved cell values and
returns the workbook as bytes, ready to stream. Value *formatting* (how a meta
field, dimension, or static column turns into a cell) lives in
`format_export_value` so it stays consistent across activities, entities, and
the future Reports builder — the on-screen `renderCellValue` is the mirror.

XLSX-only by design (see the export proposal): real Excel gives typed date and
number cells and styled headers, which the Reports feature will build on.
"""

import re
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Hard ceiling on rows per export. Generation is synchronous and in-memory, so
# this keeps a single request from exhausting memory / timing out. When a query
# would exceed it, the caller tells the user to narrow their filters. Bump this
# (or move to a background job) if NGO datasets outgrow it.
EXPORT_ROW_CAP = 25_000

_DATE_FORMAT = "dd mmm yyyy"
_DATETIME_FORMAT = "dd mmm yyyy hh:mm"
_MAX_COL_WIDTH = 60
_MIN_COL_WIDTH = 10


def format_export_value(field_type: str | None, value: Any) -> Any:
    """Turn a stored value into a cell value, mirroring the frontend renderCellValue.

    Returns native Python types where it helps Excel (date/datetime objects,
    ints/floats) and strings otherwise. Empty/missing values become "".
    """
    if value is None or value == "":
        return ""

    # Date/datetime meta fields are stored as ISO strings — parse to real
    # date objects so Excel treats them as dates (sortable, filterable).
    if field_type == "date":
        return _parse_date(value)
    if field_type == "datetime":
        return _parse_datetime(value)

    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        return ", ".join(f"{k}: {v}" for k, v in value.items())
    return value


def export_filename(name: str, ext: str = "xlsx") -> str:
    """Build a download filename: '<slug>_<YYYYMMDD>.<ext>'."""
    slug = re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_").lower() or "export"
    return f"{slug}_{datetime.now().strftime('%Y%m%d')}.{ext}"


def build_xlsx(headers: list[str], rows: list[list[Any]], sheet_name: str = "Export") -> bytes:
    """Build an XLSX workbook from headers + a row matrix and return its bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(sheet_name)

    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="EFEFEF")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([_coerce(v) for v in row])

    _apply_number_formats(ws)
    _autosize_columns(ws, headers, rows)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ── internals ───────────────────────────────────────────────────


def _coerce(value: Any) -> Any:
    """Coerce a value to something openpyxl can write.

    openpyxl rejects timezone-aware datetimes, so strip tzinfo (values are
    normalized to UTC elsewhere). Everything non-primitive falls back to str.
    """
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, (str, int, float, date)):
        return value
    return str(value)


def _apply_number_formats(ws) -> None:
    """Give date/datetime cells a readable Excel number format."""
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            val = cell.value
            if isinstance(val, datetime):
                cell.number_format = _DATETIME_FORMAT
            elif isinstance(val, date):
                cell.number_format = _DATE_FORMAT


def _autosize_columns(ws, headers: list[str], rows: list[list[Any]]) -> None:
    """Set column widths to roughly fit content, clamped to sane bounds."""
    for idx, header in enumerate(headers):
        longest = len(str(header))
        for row in rows:
            if idx < len(row) and row[idx] not in (None, ""):
                longest = max(longest, len(str(row[idx])))
        width = max(_MIN_COL_WIDTH, min(longest + 2, _MAX_COL_WIDTH))
        ws.column_dimensions[get_column_letter(idx + 1)].width = width


def _safe_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no []:*?/\\ — fall back to 'Export'."""
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", name or "").strip()
    return cleaned[:31] or "Export"


def _parse_date(value: Any):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00")).date()
        except ValueError:
            return value
    return value


def _parse_datetime(value: Any):
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
            return dt.replace(tzinfo=None) if dt.tzinfo else dt
        except ValueError:
            return value
    return value
